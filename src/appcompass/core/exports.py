"""엑셀(.xlsx) 내보내기.

보고서 Markdown은 읽기용이고, 엑셀은 **작업용**이다.
평가 점수를 정렬하거나, 기능 목록을 백로그로 옮기거나,
측정 이벤트를 개발자에게 넘길 때 쓴다.

그래서 시트를 목적별로 나눈다. 한 시트에 다 넣으면 아무 것도 못 한다.
"""

from __future__ import annotations

from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .enums import Severity
from .models import AnalysisResult, EvidenceItem

_HEADER_FILL = PatternFill("solid", fgColor="EEF1F5")
_HEADER_FONT = Font(bold=True, color="1C1C1E")
_TITLE_FONT = Font(bold=True, size=13)
_CRITICAL_FONT = Font(color="B3261E", bold=True)
_WARN_FONT = Font(color="8A6100")
_THIN = Side(style="thin", color="D5D8DC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_workbook(
    result: AnalysisResult,
    evidence: Sequence[EvidenceItem] = (),
    project_name: str = "",
    version_no: int | None = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, result, project_name, version_no)
    _sheet_scores(wb, result)
    _sheet_warnings(wb, result)
    _sheet_actions(wb, result)
    _sheet_targets(wb, result)
    _sheet_mvp(wb, result)
    _sheet_metrics(wb, result)
    _sheet_evidence(wb, evidence)
    return wb


def save_workbook(
    path: str,
    result: AnalysisResult,
    evidence: Sequence[EvidenceItem] = (),
    project_name: str = "",
    version_no: int | None = None,
) -> str:
    build_workbook(result, evidence, project_name, version_no).save(path)
    return path


# ---------------------------------------------------------------------------


def _write_table(
    ws,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    widths: Sequence[int] | None = None,
    start_row: int = 1,
) -> int:
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center")

    for r, row in enumerate(rows, start_row + 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if widths:
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(rows) + 1


def _sheet_summary(wb, result, project_name, version_no) -> None:
    ws = wb.create_sheet("요약")
    meta, diag, pivot = result.meta, result.diagnosis, result.pivot

    ws["A1"] = f"{result.idea.app_name or project_name or '기획 진단'} 진단 요약"
    ws["A1"].font = _TITLE_FONT

    rows = [
        ("프로젝트", project_name or "-"),
        ("기획 버전", f"v{version_no}" if version_no is not None else "-"),
        ("도메인", str(meta.domain_code)),
        ("생성 시각", meta.created_at.isoformat() if meta.created_at else "-"),
        ("판정 엔진", f"{meta.engine} {meta.engine_version}"),
        ("평가 정책", meta.policy_version),
        ("스키마 버전", meta.schema_version),
        ("모델", meta.model_name or "사용 안 함 (규칙 엔진 전용)"),
        ("", ""),
        ("총점", round(diag.total_score, 1)),
        ("근거 신뢰도", round(diag.overall_confidence, 2)),
        ("판단", str(pivot.decision)),
        ("근거가 충분했다면", str(pivot.would_be_decision) if pivot.would_be_decision else "-"),
        ("사유 코드", ", ".join(pivot.reason_codes) or "-"),
        ("사람 승인 필요", "예" if pivot.requires_human_approval else "아니오"),
        ("", ""),
        ("판단 근거", pivot.rationale),
    ]
    _write_table(ws, ["항목", "값"], rows, widths=[22, 110], start_row=3)


def _sheet_scores(wb, result) -> None:
    ws = wb.create_sheet("평가 점수")
    rows = [
        (
            str(d.code),
            d.label,
            d.raw_score,
            d.weight,
            round(d.normalized_score, 1),
            round(d.confidence, 2),
            d.reason,
            "; ".join(d.missing_evidence) or "-",
            d.recommended_action,
        )
        for d in result.diagnosis.dimensions
    ]
    end = _write_table(
        ws,
        ["코드", "항목", "점수(0-5)", "가중치", "환산", "신뢰도", "이유", "부족한 근거", "권장 행동"],
        rows,
        widths=[8, 20, 11, 9, 9, 9, 60, 40, 45],
    )
    # 2점 이하는 눈에 띄게
    for r in range(2, 2 + len(rows)):
        if ws.cell(row=r, column=3).value <= 2:
            ws.cell(row=r, column=3).font = _CRITICAL_FONT

    ws.cell(row=end + 1, column=1, value="합계")
    ws.cell(row=end + 1, column=1).font = _HEADER_FONT
    ws.cell(row=end + 1, column=4, value=f"=SUM(D2:D{1 + len(rows)})").font = _HEADER_FONT
    ws.cell(row=end + 1, column=5, value=f"=SUM(E2:E{1 + len(rows)})").font = _HEADER_FONT


def _sheet_warnings(wb, result) -> None:
    ws = wb.create_sheet("경고")
    order = {Severity.CRITICAL: 0, Severity.WARN: 1, Severity.INFO: 2}
    warnings = sorted(result.diagnosis.warnings, key=lambda w: order[w.severity])
    label = {Severity.CRITICAL: "치명", Severity.WARN: "주의", Severity.INFO: "참고"}
    rows = [
        (label[w.severity], str(w.code), w.field or "-", w.message, w.recommended_action)
        for w in warnings
    ]
    _write_table(
        ws, ["심각도", "코드", "대상 필드", "내용", "권장 행동"], rows,
        widths=[10, 26, 20, 70, 55],
    )
    for r, w in enumerate(warnings, 2):
        ws.cell(row=r, column=1).font = (
            _CRITICAL_FONT if w.severity == Severity.CRITICAL else _WARN_FONT
        )


def _sheet_actions(wb, result) -> None:
    ws = wb.create_sheet("할 일")
    rows: list[tuple] = []
    for a in result.next_actions:
        rows.append(("다음 행동", a, "", ""))
    for a in result.pivot.change:
        rows.append(("변경", a, "", ""))
    for a in result.pivot.remove:
        rows.append(("삭제", a, "", ""))
    for a in result.pivot.keep:
        rows.append(("유지", a, "", ""))
    for u in result.diagnosis.unknowns:
        rows.append(("언노운 확인", u, "", ""))
    _write_table(
        ws, ["구분", "내용", "담당", "기한"], rows, widths=[14, 95, 14, 14]
    )


def _sheet_targets(wb, result) -> None:
    ws = wb.create_sheet("타깃 후보")
    ws["A1"] = result.targets.recommendation_reason
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 32

    rec = result.targets.recommended_candidate_index
    rows = [
        (
            "★ 추천" if rec == i else "",
            c.name,
            c.user,
            c.payer or "-",
            c.influencer or "-",
            c.trigger_situation or "-",
            c.problem or "-",
            "\n".join(c.risks) or "-",
            "\n".join(c.validation_questions) or "-",
            c.recommended_experiment or "-",
        )
        for i, c in enumerate(result.targets.candidates)
    ]
    _write_table(
        ws,
        ["추천", "이름", "사용자", "구매자", "영향자", "발생 상황", "문제", "위험", "검증 질문", "추천 실험"],
        rows,
        widths=[8, 22, 50, 16, 20, 40, 45, 45, 45, 45],
        start_row=3,
    )


def _sheet_mvp(wb, result) -> None:
    ws = wb.create_sheet("MVP 백로그")
    m = result.mvp
    rows: list[tuple] = []
    for i, f in enumerate(m.p0_features, 1):
        rows.append((f"P0-{i:02d}", "P0", f, "핵심 문제 해결", "", "", ""))
    for i, f in enumerate(m.p1_features, 1):
        rows.append((f"P1-{i:02d}", "P1", f, "핵심 행동 완료율", "", "", ""))
    for i, f in enumerate(m.excluded_features, 1):
        rows.append((f"X-{i:02d}", "제외", f, "이번 MVP에서 만들지 않음", "-", "-", "-"))
    _write_table(
        ws,
        ["ID", "우선순위", "기능", "분류", "담당", "예상 규모", "완료 기준"],
        rows,
        widths=[10, 10, 55, 24, 12, 12, 45],
    )

    ws2 = wb.create_sheet("MVP 가설·흐름")
    rows2 = [
        ("핵심 가설", m.core_hypothesis),
        ("문제 가설", m.problem_hypothesis),
        ("행동 가설", m.behavior_hypothesis),
        ("가치 가설", m.value_hypothesis),
        ("재방문 가설", m.retention_hypothesis),
        ("수익 가설", m.revenue_hypothesis or "-"),
        ("첫 성공 경험", m.first_success_experience),
        ("", ""),
    ]
    rows2 += [(f"흐름 {i}", s) for i, s in enumerate(m.core_user_flow, 1)]
    rows2 += [("", "")] + [("위험", r) for r in m.risks]
    _write_table(ws2, ["항목", "내용"], rows2, widths=[18, 110])


def _sheet_metrics(wb, result) -> None:
    ws = wb.create_sheet("측정 이벤트")
    rows = [(name, "", "", "", "") for name in result.mvp.metrics]
    _write_table(
        ws,
        ["이벤트 이름", "발생 시점", "필수 속성", "검증 가설", "구현 여부"],
        rows,
        widths=[34, 40, 40, 18, 12],
    )
    ws.cell(row=len(rows) + 3, column=1, value="측정 이벤트가 연결되지 않은 기능은 만들지 않습니다.")
    ws.cell(row=len(rows) + 3, column=1).font = _WARN_FONT


def _sheet_evidence(wb, evidence) -> None:
    ws = wb.create_sheet("근거")
    if not evidence:
        ws["A1"] = "등록된 근거가 없습니다. 모든 항목의 신뢰도가 상한값으로 고정되어 판단을 확정할 수 없습니다."
        ws["A1"].font = _CRITICAL_FONT
        ws.column_dimensions["A"].width = 100
        return
    rows = [
        (
            str(e.evidence_type),
            e.title,
            e.summary,
            e.sample_size if e.sample_size is not None else "-",
            e.observed_at.isoformat() if e.observed_at else "-",
            ", ".join(str(s) for s in e.supports) or "-",
            ", ".join(str(c) for c in e.contradicts) or "-",
            e.source_reference or "-",
        )
        for e in evidence
    ]
    _write_table(
        ws,
        ["유형", "제목", "요약", "표본", "관측 시각", "지지 항목", "반박 항목", "출처"],
        rows,
        widths=[20, 34, 55, 9, 22, 18, 18, 34],
    )
