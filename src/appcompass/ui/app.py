"""애플리케이션 진입점.

실행:
    python -m appcompass            (src를 PYTHONPATH에 넣거나 pip install -e . 후)
    python run_app.py               (저장소 루트에서 바로)
    AppCompass.exe --selftest       (배포된 실행파일 점검)
"""

from __future__ import annotations

import sys
import traceback

from PySide6.QtWidgets import QApplication, QMessageBox

from ..services.app_service import AppService
from ..storage.db import Database
from .main_window import MainWindow
from .theme import STYLESHEET


def run_selftest(report_path: str | None = None) -> int:
    """Qt 없이 분석 파이프라인을 한 번 돌려 본다.

    실행파일로 묶은 뒤 스키마 파일이 빠지면 모든 분석이 FAILED_SCHEMA가 되는데,
    GUI만으로는 창이 뜨는 것까지만 확인되고 그 사실을 놓친다.
    릴리즈 전에 이 모드로 스키마 로딩·채점·판정·보고서 생성을 한 번에 확인한다.

    성공하면 0, 실패하면 1을 반환하고 결과를 파일에 남긴다.
    """
    import tempfile
    from datetime import datetime, timezone
    from pathlib import Path

    from ..core.enums import DomainCode
    from ..core.exports import save_workbook
    from ..core.models import AnalysisResult, IdeaStructure
    from ..core.pipeline import run_analysis
    from ..core.policy import EvaluationPolicy
    from ..core.report import render_html, render_markdown
    from ..core.schema import SCHEMA_SEARCH_PATHS, validate_analysis_result
    from ..core.techspec import render_techspec

    lines: list[str] = []
    ok = True

    def check(label: str, condition: bool, detail: str = "") -> None:
        nonlocal ok
        if not condition:
            ok = False
        lines.append(f"[{'OK ' if condition else 'FAIL'}] {label}{'  ' + detail if detail else ''}")

    lines.append(f"frozen={getattr(sys, 'frozen', False)}")
    lines.append(f"schema search paths={[str(p) for p in SCHEMA_SEARCH_PATHS]}")

    try:
        idea = IdeaStructure(
            app_name="자체 점검",
            target_user=(
                "AI 코딩 도구로 처음 앱을 만들다가 API·토큰 용어를 몰라 "
                "작업이 자주 중단되는 비개발자"
            ),
            payer="사용자 본인",
            problem_situation=(
                "프로젝트를 만들다가 오류 메시지에 모르는 용어가 나오면 매번 중단되고, "
                "무엇을 다시 물어야 할지 몰라 헤매다 그날 작업을 포기한다."
            ),
            current_solution="검색하거나 AI에게 되묻는다",
            current_solution_problem="검색 결과가 내 상황과 맞지 않아 시간이 오래 걸리고 다시 막힌다",
            core_action="막힌 단계를 골라 실제 상황형 미션 하나를 3분 안에 완료한다",
            expected_result="학습 후 작업 재개율이 50% 이상이 된다",
            first_success="가입 없이 첫 3분 미션 하나를 완료한다",
            retention_reason="매일 다음 단계 용어가 열리고 틀린 개념이 복습으로 돌아온다",
            distribution_channel="바이브코딩 커뮤니티와 오픈채팅",
        )
        result = run_analysis(
            idea,
            domain_code=DomainCode.VIBEQUEST,
            policy=EvaluationPolicy(),
            now=datetime(2026, 1, 1, tzinfo=timezone.utc),
        )
        check("분석 실행", True)

        payload = result.to_dict()
        validate_analysis_result(payload)
        check("JSON Schema 검증", True)

        check(
            "10개 항목 채점",
            len(payload["diagnosis"]["dimensions"]) == 10,
            f"{len(payload['diagnosis']['dimensions'])}개",
        )
        check(
            "근거 없음 → HOLD",
            payload["pivot"]["decision"] == "HOLD",
            payload["pivot"]["decision"],
        )
        check(
            "would_be_decision 존재",
            bool(payload["pivot"]["would_be_decision"]),
            str(payload["pivot"]["would_be_decision"]),
        )
        check("도메인 제외 기능 적용", "실시간 PvP" in " ".join(payload["mvp"]["excluded_features"]))

        md = render_markdown(result, project_name="자체 점검")
        html_doc = render_html(result, project_name="자체 점검")
        check("Markdown 보고서", "## 판단" in md, f"{len(md)}자")
        check("HTML 보고서", html_doc.startswith("<!DOCTYPE html>"), f"{len(html_doc)}자")

        spec = render_techspec(result, project_name="자체 점검")
        check(
            "TECHSPEC 기술 명세",
            "## 3. MVP 범위" in spec and "#### P0-01" in spec,
            f"{len(spec)}자",
        )

        # 엑셀은 외부 라이브러리(openpyxl)를 쓴다.
        # 실행파일에 번들되지 않으면 여기서만 드러난다.
        restored = AnalysisResult.from_dict(payload)
        check("분석 결과 직렬화 왕복", restored.to_dict() == payload)

        # ignore_cleanup_errors: 임시 파일을 백신·인덱서가 잡고 있으면 삭제가 실패한다.
        # 점검 대상은 엑셀 생성이지 임시 폴더 정리가 아니다.
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            xlsx_path = Path(tmp) / "selftest.xlsx"
            save_workbook(str(xlsx_path), restored, project_name="자체 점검")
            size = xlsx_path.stat().st_size if xlsx_path.exists() else 0
            check("엑셀 생성", size > 0, f"{size:,}바이트")

            from openpyxl import load_workbook

            # Windows에서는 파일을 닫지 않으면 임시 폴더를 지울 수 없다.
            check_wb = load_workbook(xlsx_path)
            try:
                sheets = check_wb.sheetnames
            finally:
                check_wb.close()
            check(
                "엑셀 시트 구성",
                "MVP 백로그" in sheets and "평가 점수" in sheets,
                f"{len(sheets)}개 시트",
            )

        db = Database(url="sqlite:///:memory:")
        AppService(db)
        db.dispose()
        check("DB 스키마 생성", True)

        lines.append("")
        lines.append(f"총점 {payload['diagnosis']['total_score']:.1f} / 100")
        lines.append(f"신뢰도 {payload['diagnosis']['overall_confidence']:.2f}")
        lines.append(f"판단 {payload['pivot']['decision']}")
    except Exception as exc:  # noqa: BLE001
        ok = False
        lines.append(f"[FAIL] 예외 발생: {type(exc).__name__}: {exc}")
        lines.append(traceback.format_exc())

    lines.append("")
    lines.append("RESULT: PASS" if ok else "RESULT: FAIL")
    text = "\n".join(lines)

    print(text)
    target = Path(report_path or "selftest.txt")
    try:
        target.write_text(text, encoding="utf-8")
    except OSError:
        pass
    return 0 if ok else 1


def main(argv: list[str] | None = None) -> int:
    argv = argv if argv is not None else sys.argv

    if "--selftest" in argv:
        index = argv.index("--selftest")
        out = argv[index + 1] if len(argv) > index + 1 else None
        return run_selftest(out)

    app = QApplication(argv)
    app.setApplicationName("AppCompass AI")
    app.setOrganizationName("AppCompass")
    app.setStyleSheet(STYLESHEET)

    try:
        db = Database()
        service = AppService(db)
    except Exception as exc:  # noqa: BLE001 - 기동 실패는 사용자에게 보여준다
        QMessageBox.critical(
            None,
            "시작 실패",
            f"데이터베이스를 열 수 없습니다.\n\n{type(exc).__name__}: {exc}\n\n"
            f"{traceback.format_exc(limit=3)}",
        )
        return 1

    window = MainWindow(service)
    window.show()
    try:
        return app.exec()
    finally:
        db.dispose()


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
