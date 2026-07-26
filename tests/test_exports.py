"""TECHSPEC 생성과 엑셀 내보내기 테스트."""

from __future__ import annotations

from datetime import datetime, timezone

import pytest

from appcompass.core.enums import (
    REPORT_FORMAT_SUFFIX,
    STORED_REPORT_FORMATS,
    DimensionCode,
    DomainCode,
    EvidenceType,
    ReportFormat,
)
from appcompass.core.exports import build_workbook
from appcompass.core.models import AnalysisResult, EvidenceItem
from appcompass.core.pipeline import run_analysis
from appcompass.core.policy import EvaluationPolicy
from appcompass.core.techspec import TODO, render_techspec

from conftest import fixture_idea, fixture_raw

FIXED_NOW = datetime(2026, 1, 1, tzinfo=timezone.utc)


def analyze(path, domain, evidence=()):
    return run_analysis(
        fixture_idea(path),
        domain_code=domain,
        policy=EvaluationPolicy(),
        evidence=evidence,
        now=FIXED_NOW,
    )


# ==========================================================================
# TECHSPEC
# ==========================================================================


def test_techspec_has_implementable_sections():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    doc = render_techspec(result, project_name="examath", version_no=3)
    for section in (
        "## 1. 제품 정의",
        "## 2. 이 MVP가 검증할 가설",
        "## 3. MVP 범위",
        "## 4. 화면과 흐름",
        "## 5. 측정 이벤트",
        "## 6. 데이터 모델",
        "## 8. 위험과 대응",
        "## 11. 구현 순서",
        "## 12. 완료의 정의",
    ):
        assert section in doc, f"TECHSPEC에 {section} 이 없습니다."


def test_techspec_gives_each_feature_an_id_and_spec_block():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    doc = render_techspec(result, project_name="examath")
    assert "#### P0-01" in doc
    # 각 기능마다 구현에 필요한 항목이 붙어야 한다
    for row in ("| 목적 |", "| 입력 |", "| 출력 |", "| 실패·예외 |", "| 완료 기준 |"):
        assert row in doc


def test_techspec_marks_underivable_fields_instead_of_inventing():
    """분석에서 알 수 없는 항목은 지어내지 않고 [결정 필요]로 남긴다."""
    result = analyze("examath/broad_child.json", DomainCode.EXAMATH)
    doc = render_techspec(result, project_name="examath")
    assert TODO in doc
    assert doc.count(TODO) >= 10, "모르는 항목을 임의로 채운 것으로 보입니다."


def test_techspec_warns_when_decision_is_hold():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    assert result.pivot.decision.value == "HOLD"
    doc = render_techspec(result, project_name="examath")
    assert "구현 착수 전 확인" in doc
    assert "HOLD" in doc


def test_techspec_embeds_domain_constraints():
    em = render_techspec(analyze("examath/refined_target.json", DomainCode.EXAMATH))
    assert "구체물" in em and "그림" in em and "숫자" in em
    assert "익명" in em or "가명" in em

    vq = render_techspec(analyze("vibequest/refined_target.json", DomainCode.VIBEQUEST))
    assert "실제 작업 상황" in vq
    assert "프로젝트 코드를 저장하지 않는다" in vq


def test_techspec_lists_excluded_features():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    doc = render_techspec(result)
    assert "이번에는 만들지 않는다" in doc
    for banned in ("광고", "가챠", "실시간 랭킹"):
        assert banned in doc


def test_techspec_records_versions():
    result = analyze("vibequest/refined_target.json", DomainCode.VIBEQUEST)
    doc = render_techspec(result, project_name="VibeQuest", version_no=2)
    assert "RULE_ENGINE" in doc
    assert EvaluationPolicy().version in doc
    assert "techspec-0.1.0" in doc


def test_techspec_is_deterministic():
    a = render_techspec(analyze("vibequest/refined_target.json", DomainCode.VIBEQUEST))
    b = render_techspec(analyze("vibequest/refined_target.json", DomainCode.VIBEQUEST))
    assert a == b


# ==========================================================================
# 직렬화 왕복
# ==========================================================================


def test_analysis_result_roundtrip():
    """엑셀은 저장된 JSON에서 만들어지므로 왕복이 손실 없어야 한다."""
    original = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    restored = AnalysisResult.from_dict(original.to_dict())
    assert restored.to_dict() == original.to_dict()


# ==========================================================================
# 엑셀
# ==========================================================================


def test_workbook_has_purpose_separated_sheets():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    wb = build_workbook(result, project_name="examath", version_no=1)
    for name in (
        "요약", "평가 점수", "경고", "할 일",
        "타깃 후보", "MVP 백로그", "MVP 가설·흐름", "측정 이벤트", "근거",
    ):
        assert name in wb.sheetnames, f"엑셀에 '{name}' 시트가 없습니다."


def test_workbook_scores_sheet_has_ten_rows():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    ws = build_workbook(result)["평가 점수"]
    codes = [ws.cell(row=r, column=1).value for r in range(2, 12)]
    assert codes == [f"D{i:02d}" for i in range(1, 11)]


def test_workbook_backlog_assigns_ids():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    ws = build_workbook(result)["MVP 백로그"]
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert any(str(i).startswith("P0-") for i in ids)
    assert any(str(i).startswith("X-") for i in ids)


def test_workbook_states_missing_evidence_clearly():
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH)
    ws = build_workbook(result, evidence=[])["근거"]
    assert "등록된 근거가 없습니다" in str(ws["A1"].value)


def test_workbook_includes_evidence_when_present():
    evidence = [
        EvidenceItem(
            id="e1",
            evidence_type=EvidenceType.USER_INTERVIEW,
            title="학부모 인터뷰 8명",
            sample_size=8,
            supports=(DimensionCode.D01,),
        )
    ]
    result = analyze("examath/refined_target.json", DomainCode.EXAMATH, evidence)
    ws = build_workbook(result, evidence=evidence)["근거"]
    assert ws.cell(row=2, column=2).value == "학부모 인터뷰 8명"


# ==========================================================================
# 서비스 계층 통합
# ==========================================================================


def test_analysis_stores_all_text_formats(service):
    project = service.create_project("내보내기", domain_code=DomainCode.EXAMATH)
    version = service.create_version(
        project.id,
        fixture_raw("examath/refined_target.json"),
        fixture_idea("examath/refined_target.json"),
    )
    service.approve_structure(version.id)
    run = service.run_analysis(version.id)

    formats = {r.format for r in service.get_reports(run.id)}
    for fmt in STORED_REPORT_FORMATS:
        assert str(fmt) in formats, f"{fmt} 보고서가 저장되지 않았습니다."


@pytest.mark.parametrize(
    "fmt", [ReportFormat.MARKDOWN, ReportFormat.HTML, ReportFormat.TECHSPEC, ReportFormat.XLSX]
)
def test_export_every_format(service, tmp_path, fmt):
    project = service.create_project("내보내기", domain_code=DomainCode.VIBEQUEST)
    version = service.create_version(
        project.id,
        fixture_raw("vibequest/refined_target.json"),
        fixture_idea("vibequest/refined_target.json"),
    )
    service.approve_structure(version.id)
    run = service.run_analysis(version.id)

    target = tmp_path / f"out{REPORT_FORMAT_SUFFIX[fmt]}"
    service.export_report(run.id, fmt, str(target))
    assert target.exists()
    assert target.stat().st_size > 0

    if fmt == ReportFormat.XLSX:
        from openpyxl import load_workbook

        wb = load_workbook(target)
        assert "MVP 백로그" in wb.sheetnames
    elif fmt == ReportFormat.TECHSPEC:
        assert "기술 명세" in target.read_text(encoding="utf-8")

    assert "REPORT_EXPORTED" in [a.action for a in service.list_audit_logs()]
